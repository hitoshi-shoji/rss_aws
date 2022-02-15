import sqlite3
import openpyxl
import datetime
import os
import settings
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

rss_new_url = settings.rss_new_url
rss_blog = settings.rss_blog
rss_blog_url = settings.rss_blog_url
dbname = settings.sqlite_dbname
now = datetime.datetime.now()
filename = settings.output_path + 'aws-rss_' + now.strftime('%Y%m%d.%H%M%S') + '.xlsx'

def main():
    rssawsnew_excel()
    if rss_blog == "true":
        rssawsblog_excel()


def rssawsnew_excel():
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = 'Work(Whats New)'


    border = Border(top=Side(style='thin', color='000000'), 
        bottom=Side(style='thin', color='000000'), 
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000')
    )

    fill = PatternFill(patternType='solid',fgColor='00FFFF')

    sheet1['A2'].value = 'No.'
    sheet1['B2'].value = '投稿日'
    sheet1['C2'].value = 'サービス'
    sheet1['D2'].value = 'アップデートタイトル'
    sheet1['E2'].value = '概要'
    sheet1['F2'].value = '調査対象'
    sheet1.auto_filter.ref = "A2:F2"

    for abc in ['A','B','C','D','E','F']:
        sheet1[abc + '2'].font = Font(name='Meiryo',size='10',bold=True)
        sheet1[abc + '2'].border = border 
        sheet1[abc + '2'].alignment = Alignment(horizontal='center', vertical='center') 
        sheet1[abc + '2'].fill = fill 

    db = sqlite3.connect(dbname)
    cur = db.cursor()
    cur.execute('SELECT no,published,category,title,url,summary FROM rss_new_table order by published desc')
    for i, row in enumerate (cur):
        sheet1['A' + str(i+3)].value =  (i+1)
        sheet1['B' + str(i+3)].value = row[1] 
        sheet1['C' + str(i+3)].value = row[2] 
        sheet1['D' + str(i+3)].value = row[3] 
        sheet1['D' + str(i+3)].hyperlink = row[4]
        sheet1['D' + str(i+3)].style = "Hyperlink"
        sheet1['E' + str(i+3)].value = row[5] 

    cur.close()

    for rows0 in sheet1['A3':'A1000']:
        for cell0 in rows0:
            cell0.font = Font(name='Meiryo',size='9')
            cell0.border = border
            cell0.alignment = Alignment(horizontal='center', vertical='center')

    for rows1 in sheet1['B3':'C1000']:
        for cell1 in rows1:
            cell1.font = Font(name='Meiryo',size='9')
            cell1.border = border
            cell1.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    for rows2 in sheet1['D3':'D1000']:
        for cell2 in rows2:
            cell2.font = Font(name='Meiryo',size='9',color='0000FF')
            cell2.border = border
            cell2.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

    for rows3 in sheet1['E3':'E1000']:
        for cell3 in rows3:
            cell3.font = Font(name='Meiryo',size='9')
            cell3.border = border
            cell3.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

    for rows4 in sheet1['F3':'F1000']:
        for cell4 in rows4:
            cell4.font = Font(name='Meiryo',size='9')
            cell4.border = border
            cell4.alignment = Alignment(horizontal='center', vertical='center')

    sheet1.column_dimensions['A'].width = 4
    sheet1.column_dimensions['B'].width = 12
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 55 
    sheet1.column_dimensions['E'].width = 82 
    sheet1.column_dimensions['F'].width = 12 
    sheet1.freeze_panes = 'A3'

    workbook.save(filename)

def rssawsblog_excel():
    workbook = openpyxl.load_workbook(filename)
    workbook.create_sheet( title="Work(AWS Blog)" )
    sheet2 = workbook['Work(AWS Blog)']

    border = Border(top=Side(style='thin', color='000000'), 
        bottom=Side(style='thin', color='000000'), 
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000')
    )

    fill = PatternFill(patternType='solid',fgColor='00FFFF')

    sheet2['A2'].value = 'No.'
    sheet2['B2'].value = '投稿日'
    sheet2['C2'].value = 'サービス'
    sheet2['D2'].value = 'アップデートタイトル'
    sheet2['E2'].value = '概要'
    sheet2['F2'].value = '調査対象'
    sheet2.auto_filter.ref = "A2:F2"

    for abc in ['A','B','C','D','E','F']:
        sheet2[abc + '2'].font = Font(name='Meiryo',size='10',bold=True)
        sheet2[abc + '2'].border = border 
        sheet2[abc + '2'].alignment = Alignment(horizontal='center', vertical='center') 
        sheet2[abc + '2'].fill = fill 

    db = sqlite3.connect(dbname)
    cur = db.cursor()
    cur.execute('SELECT no,published,category,title,url,summary FROM rss_blog_table order by published desc')
    for i, row in enumerate (cur):
        sheet2['A' + str(i+3)].value =  (i+1)
        sheet2['B' + str(i+3)].value = row[1] 
        sheet2['C' + str(i+3)].value = row[2] 
        sheet2['D' + str(i+3)].value = row[3] 
        sheet2['D' + str(i+3)].hyperlink = row[4]
        sheet2['D' + str(i+3)].style = "Hyperlink"
        sheet2['E' + str(i+3)].value = row[5] 

    cur.close()

    for rows0 in sheet2['A3':'A1000']:
        for cell0 in rows0:
            cell0.font = Font(name='Meiryo',size='9')
            cell0.border = border
            cell0.alignment = Alignment(horizontal='center', vertical='center')

    for rows1 in sheet2['B3':'C1000']:
        for cell1 in rows1:
            cell1.font = Font(name='Meiryo',size='9')
            cell1.border = border
            cell1.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    for rows2 in sheet2['D3':'D1000']:
        for cell2 in rows2:
            cell2.font = Font(name='Meiryo',size='9',color='0000FF')
            cell2.border = border
            cell2.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

    for rows3 in sheet2['E3':'E1000']:
        for cell3 in rows3:
            cell3.font = Font(name='Meiryo',size='9')
            cell3.border = border
            cell3.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

    for rows4 in sheet2['F3':'F1000']:
        for cell4 in rows4:
            cell4.font = Font(name='Meiryo',size='9')
            cell4.border = border
            cell4.alignment = Alignment(horizontal='center', vertical='center')

    sheet2.column_dimensions['A'].width = 4
    sheet2.column_dimensions['B'].width = 12
    sheet2.column_dimensions['C'].width = 20
    sheet2.column_dimensions['D'].width = 55 
    sheet2.column_dimensions['E'].width = 82 
    sheet2.column_dimensions['F'].width = 12 
    sheet2.freeze_panes = 'A3'

    workbook.save(filename)

if __name__ == "__main__":
    main()


